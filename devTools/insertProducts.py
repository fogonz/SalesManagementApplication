import openpyxl
import mysql.connector
import warnings
from mysql.connector import Error
import logging
from datetime import datetime

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('procesamiento_productos.log'),
        logging.StreamHandler()
    ]
)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ProcesadorProductos:
    def __init__(self, db_config, xlsx_path, sheet_name="STOCK FELI"):
        self.db_config = db_config
        self.xlsx_path = xlsx_path
        self.sheet_name = sheet_name
        self.conn = None
        self.cursor = None
        self.wb = None
        self.ws = None
        
    def __enter__(self):
        """Context manager para manejo automático de recursos"""
        try:
            # Cargar archivo Excel
            logging.info(f"Cargando archivo: {self.xlsx_path}")
            self.wb = openpyxl.load_workbook(self.xlsx_path)
            
            if self.sheet_name not in self.wb.sheetnames:
                available_sheets = ", ".join(self.wb.sheetnames)
                raise ValueError(f"Hoja '{self.sheet_name}' no encontrada. Hojas disponibles: {available_sheets}")
            
            self.ws = self.wb[self.sheet_name]
            logging.info(f"Hoja '{self.sheet_name}' cargada exitosamente")
            
            # Conectar a base de datos
            logging.info("Conectando a la base de datos...")
            self.conn = mysql.connector.connect(**self.db_config)
            self.cursor = self.conn.cursor()
            logging.info("Conexión establecida exitosamente")
            
            return self
            
        except Exception as e:
            logging.error(f"Error en inicialización: {e}")
            self.__exit__(None, None, None)
            raise
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Cleanup automático de recursos"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            if exc_type is None:
                self.conn.commit()
                logging.info("Transacción confirmada")
            else:
                self.conn.rollback()
                logging.warning("Transacción revertida debido a errores")
            self.conn.close()
        if self.wb:
            self.wb.close()
        logging.info("Recursos liberados")
    
    @staticmethod
    def limpiar_precio(val):
        """
        Convierte valores de precio a float, manejando diferentes formatos
        """
        if val is None or val == "":
            return 0.0
        
        try:
            # Convertir a string y limpiar
            val_str = str(val).strip()
            
            # Remover símbolos de moneda y espacios
            val_str = val_str.replace('$', '').replace('€', '').replace(' ', '')
            
            # Manejar separadores decimales (coma y punto)
            if ',' in val_str and '.' in val_str:
                # Formato: 1,234.56 (punto decimal, coma miles)
                val_str = val_str.replace(',', '')
            elif ',' in val_str and val_str.count(',') == 1:
                # Formato: 1234,56 (coma decimal)
                val_str = val_str.replace(',', '.')
            
            return float(val_str)
            
        except (ValueError, TypeError) as e:
            logging.warning(f"No se pudo convertir '{val}' a float: {e}")
            return 0.0
    
    @staticmethod
    def limpiar_entero(val):
        """
        Convierte valores a entero
        """
        if val is None or val == "":
            return 0
        
        try:
            # Primero convertir a float para manejar decimales, luego a int
            return int(float(str(val).replace(',', '.')))
        except (ValueError, TypeError) as e:
            logging.warning(f"No se pudo convertir '{val}' a entero: {e}")
            return 0
    
    def validar_fila(self, row, row_num):
        """
        Valida que una fila tenga datos mínimos requeridos
        """
        if not row or len(row) < 6:
            logging.warning(f"Fila {row_num}: Datos insuficientes (menos de 6 columnas)")
            return False
        
        # Verificar que al menos el nombre del producto no esté vacío
        if not row[0] or str(row[0]).strip() == "":
            logging.warning(f"Fila {row_num}: Nombre de producto vacío")
            return False
        
        return True
    
    def construir_tipo_producto(self, nombre, largo, ancho):
        """
        Construye el tipo de producto concatenando nombre, largo y ancho
        """
        # Limpiar valores None o vacíos
        nombre = str(nombre).strip() if nombre else ""
        largo = str(largo).strip() if largo else ""
        ancho = str(ancho).strip() if ancho else ""
        
        # Construir tipo_producto eliminando espacios extra
        tipo_producto = f"{nombre} {largo} {ancho}".strip()
        
        # Limpiar espacios múltiples
        while "  " in tipo_producto:
            tipo_producto = tipo_producto.replace("  ", " ")
        
        return tipo_producto
    
    def verificar_duplicado(self, tipo_producto):
        """
        Verifica si ya existe un producto con el mismo tipo_producto
        """
        try:
            check_query = "SELECT id FROM productos WHERE tipo_producto = %s"
            self.cursor.execute(check_query, (tipo_producto,))
            result = self.cursor.fetchone()
            return result is not None
        except Error as e:
            logging.error(f"Error verificando duplicado: {e}")
            return False
    
    def procesar_archivo(self):
        """
        Procesa el archivo Excel e inserta/actualiza datos en MySQL.
        Siempre actualiza los duplicados, nunca los omite ni borra.
        """
        logging.info("Iniciando procesamiento del archivo...")

        # Estadísticas
        total_filas = 0
        procesadas = 0
        errores = 0
        duplicados = 0
        actualizados = 0

        # SQL queries
        insert_query = """
            INSERT INTO productos (tipo_producto, precio_venta_unitario, costo_unitario, cantidad_inicial)
            VALUES (%s, %s, %s, %s)
        """

        update_query = """
            UPDATE productos 
            SET precio_venta_unitario = %s, costo_unitario = %s, cantidad_inicial = %s
            WHERE tipo_producto = %s
        """

        try:
            # Obtener información sobre las columnas
            max_row = self.ws.max_row
            max_col = self.ws.max_column
            logging.info(f"Archivo tiene {max_row} filas y {max_col} columnas")

            # Mostrar encabezados (primera fila)
            headers = [cell.value for cell in self.ws[1]]
            logging.info(f"Encabezados detectados: {headers}")

            # Procesar filas (empezando desde la fila 2)
            for row_num, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
                total_filas += 1

                # Validar fila
                if not self.validar_fila(row, row_num):
                    errores += 1
                    continue

                try:
                    # Extraer datos
                    nombre = row[0]
                    largo = row[1]
                    ancho = row[2]
                    pv = self.limpiar_precio(row[3])
                    cu = self.limpiar_precio(row[4])
                    cantidad_inicial = self.limpiar_entero(row[5])

                    # Construir tipo_producto
                    tipo_producto = self.construir_tipo_producto(nombre, largo, ancho)

                    if not tipo_producto:
                        logging.warning(f"Fila {row_num}: tipo_producto vacío después de limpieza")
                        errores += 1
                        continue

                    logging.info(f"Procesando fila {row_num}: {tipo_producto}")

                    # Verificar duplicados
                    es_duplicado = self.verificar_duplicado(tipo_producto)

                    if es_duplicado:
                        duplicados += 1
                        # Siempre actualizar duplicados
                        self.cursor.execute(update_query, (pv, cu, cantidad_inicial, tipo_producto))
                        actualizados += 1
                        logging.info(f"Fila {row_num}: Producto actualizado")
                    else:
                        # Insertar nuevo registro
                        self.cursor.execute(insert_query, (tipo_producto, pv, cu, cantidad_inicial))
                        procesadas += 1
                        logging.info(f"Fila {row_num}: Producto insertado")

                except Error as e:
                    errores += 1
                    logging.error(f"Error MySQL en fila {row_num}: {e}")
                    logging.error(f"Datos de la fila: {row}")

                except Exception as e:
                    errores += 1
                    logging.error(f"Error general en fila {row_num}: {e}")
                    logging.error(f"Datos de la fila: {row}")

            # Mostrar estadísticas finales
            logging.info("=" * 50)
            logging.info("RESUMEN DEL PROCESAMIENTO")
            logging.info("=" * 50)
            logging.info(f"Total de filas procesadas: {total_filas}")
            logging.info(f"Productos insertados: {procesadas}")
            logging.info(f"Productos actualizados: {actualizados}")
            logging.info(f"Duplicados encontrados: {duplicados}")
            logging.info(f"Errores: {errores}")
            logging.info(f"Tasa de éxito: {((procesadas + actualizados) / total_filas * 100):.2f}%")

        except Exception as e:
            logging.error(f"Error crítico durante el procesamiento: {e}")
            raise

def main():
    # Configuración de la base de datos
    db_config = {
        'user': 'root',
        'password': '765123',
        'host': 'localhost',
        'database': 'db_schema_nuevo'
    }

    # Ruta al archivo xlsx
    xlsx_path = './Maderera Movimientos.xlsx'

    try:
        # Usar context manager para manejo automático de recursos
        with ProcesadorProductos(db_config, xlsx_path, "STOCK FELI") as procesador:
            # Procesar archivo (siempre actualiza duplicados)
            procesador.procesar_archivo()

    except Exception as e:
        logging.error(f"Error en el proceso principal: {e}")
        print(f"Error: {e}")

if __name__ == "__main__":
    main()